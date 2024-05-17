from pathlib import Path
from openpyxl import load_workbook
from unidecode import unidecode
import pandas as pd
import re

def subistitute_words(text: str, maping: dict[str, str]) -> str:
    """Replace a text for the value associated with it in the dicionary.

    Args:
        text (str): A generical word.
        maping (dict[str, str]): A dictionary that contains a key and a value.

    Returns:
        str: .
    """
    pattern = r'\b(?:{})\b'.format('|'.join(map(re.escape, maping.keys())))
    corresponding = re.sub(pattern, lambda match: maping[match.group(0)], text)
    return corresponding

def normalized(text: str) -> str:
    """Remove all accents from words

    Args:
        text (str): A text will be normalized

    Returns:
        str: The corresponding text
    """
    text_normalized = unidecode(text)
    return text_normalized

def get_map(file_path: Path, from_column: str, to_column: str) -> dict[str, str]:

    """Creates a dicionary using a "FROM-TO" Excel file.

    Args:
        file_path (Path): Path of the file that the words will be changed.
        from_column (str): Name of the FROM column, this column contains the word will be
        changed.
        to_column (str): Name of the TO column, this column contains the word that 
        replaced the old one.

    Returns:
        dict[str, str]: A dictionary where the key is the word in the 'FROM' column
        and the value is the word in the 'TO' column
    """

    document = pd.read_excel(file_path)
    file_dic = dict(zip(document[from_column], document[to_column]))
    return file_dic

def excel_substitution(file_path: Path, column_name: str, from_to_path: Path, 
                       from_column: str, to_column: str, normalize:bool = False) -> None:
    
    """Goes through the specified column of the entire file replacing all words that 
    need to be changed.

    Args:
        file_path (Path): Path of the file that the words will be changed.
        column_name (str): Name of the column you need to change.
        from_to_path (Path): Path of the FROM-TO file.
        from_column (str): Name of the FROM column, this column contains the word will be
        changed.
        to_column (str): Name of the TO column, this column contains the word that 
        replaced the old one.
        normalize (bool, optional): If you need to normalize the words, normalize will be 
        True, else will be False. Defaults to False.
    """
    
    wb = load_workbook(file_path)
    sheet = wb.active
    first_row = 1
    last_row = sheet.max_row

    target_column = None
    for header in sheet[1]:
        if header.value == column_name:
            target_column = header.column
            break

    if target_column != None:
        for i in range(first_row+1, last_row+1):
            cell = str(sheet.cell(i, column=target_column).value)
            sheet.cell(i, column=target_column).value = subistitute_words(
                cell, get_map(from_to_path, from_column, to_column))
            if normalize:
                new_cell= str(sheet.cell(i, column=target_column).value)
                sheet.cell(i, column=target_column).value = normalized(new_cell)

    wb.save(file_path)