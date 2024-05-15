from pathlib import Path

from openpyxl import load_workbook
import os

from src.functions import subistitute_words, normalized, get_map

import time

def main(data_path):

    start_time = time.time()
    counter = 0

    for folder in os.listdir(data_path):
        path = data_path / folder

        if path.is_dir():
            all_files = list(path.glob("*xlsx"))

            for file in all_files:
                print(file)
                wb = load_workbook(file)
                sheet = wb.active
                first_row = 1
                last_row = sheet.max_row
                total_row = last_row - first_row
                counter += total_row

                translated = "Translated"

                target_column = None

                for header in sheet[1]:
                    if header.value == translated:
                        target_column = header.column
                        break
            
                if target_column != None:
                    for i in range(first_row+1, last_row+1):
                        cell = str(sheet.cell(i, column=target_column).value)
                        sheet.cell(i, column=target_column).value = subistitute_words(cell, get_map(file_path))
                        new_cell= str(sheet.cell(i, column=target_column).value)
                        # sheet.cell(i, column=target_column).value = normalized(new_cell)
                        
                    
                wb.save(file)

    end_time = time.time()
    total_time = start_time - end_time

    print(f"Levou {total_time} segundos para a execução e um total de {counter} linhas")




data_path = Path(__file__).parent / "data"

file_path = Path(__file__).parent / "Interface translations Chinese - document control, training management and archive.xlsx"

if __name__ == "__main__":

    main(data_path)

