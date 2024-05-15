
from pathlib import Path

from openpyxl import load_workbook
import pandas as pd
import re
import os
from unidecode import unidecode


teste_path = Path(__file__).parent / "teste.xlsx"

file_path = Path(__file__).parent / "Interface translations Chinese - document control, training management and archive.xlsx"

document = pd.read_excel(file_path)

file_dic = dict(zip(document["Current translation"], document["Correct translation"]))


def substituir_palavras_completas(text: str, maping: dict[str, str]) -> str:
    padrao = r'\b(?:{})\b'.format('|'.join(map(re.escape, maping.keys())))
    substituido = re.sub(padrao, lambda match: maping[match.group(0)], text)
    return substituido


def normalized(text: str) -> str:
    text_normalized = unidecode(text)
    return text_normalized



wb = load_workbook(teste_path)
sheet = wb.active
first_row = 1
last_row = sheet.max_row

translated = "Translated"

target_column = None

for header in sheet[1]:
    if header.value == translated:
        target_column = header.column
        break

if target_column != None:
    for i in range(first_row+1, last_row+1):
        cell = str(sheet.cell(i, column=target_column).value)
        print(cell)
        sheet.cell(i, column=target_column).value = substituir_palavras_completas(cell, file_dic)
        new_cell = str(sheet.cell(i, column=target_column).value)
        print(new_cell)
        # sheet.cell(i, column=target_column).value = normalized(new_cell)
        

wb.save(teste_path)
